from __future__ import annotations

import dataclasses
from decimal import Decimal
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.schemas.extraction import ExtractionResult
from backend.services.template_service import filter_extraction_by_template

# ── Styling constants ──────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill("solid", fgColor="2D5B8A")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_BOLD_FONT = Font(bold=True)
_MONEY_FMT = "#,##0.00"
_RED_FONT = Font(color="FF0000")

# Column name fragments that suggest a monetary value
_MONETARY_KEYS = ("amount", "price", "total", "base", "imponible", "iva", "irpf", "cuota")


# ── Helpers ────────────────────────────────────────────────────────────────────

def _is_monetary_value(value: Any) -> bool:
    """Return True if value is Decimal, float, or int (non-bool)."""
    if isinstance(value, bool):
        return False
    return isinstance(value, (Decimal, float, int))


def _is_monetary_column(name: str) -> bool:
    """Return True if the column name suggests a monetary value."""
    low = name.lower()
    return any(k in low for k in _MONETARY_KEYS)


def _to_numeric(value: Any) -> float | Any:
    """Convert Decimal to float for Excel numeric storage; return others unchanged."""
    if isinstance(value, Decimal):
        return float(value)
    return value


def _write_header_row(ws: Any, headers: list[str]) -> None:
    """Write a bold blue header row at row 1."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autofit_columns(ws: Any, min_width: int = 10, max_width: int = 50) -> None:
    """Set column widths based on max content length per column."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_len + 2))


# ── Sheet builders ─────────────────────────────────────────────────────────────

def _build_summary_sheet(
    ws: Any,
    result: ExtractionResult,
    template_fields: list[dict] | None,
) -> None:
    """Populate the Summary sheet."""
    # Determine field/value pairs
    if template_fields:
        filtered: dict[str, Any] = filter_extraction_by_template(result, template_fields)
        rows: list[tuple[str, Any]] = list(filtered.items())
    else:
        # Build from all anchor fields using dataclasses.asdict
        anchor_dict = dataclasses.asdict(result.anchor)
        rows = [(k, v) for k, v in anchor_dict.items()]

    _write_header_row(ws, ["Field", "Value"])
    ws.freeze_panes = "A2"

    for row_idx, (field_name, value) in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=field_name)
        if _is_monetary_value(value):
            cell = ws.cell(row=row_idx, column=2, value=_to_numeric(value))
            cell.number_format = _MONEY_FMT
        else:
            # Convert Decimal-looking strings or leave as-is
            display_value = str(value) if value is not None else ""
            ws.cell(row=row_idx, column=2, value=display_value if display_value else None)

    _autofit_columns(ws)


def _build_line_items_sheet(ws: Any, result: ExtractionResult) -> None:
    """Populate the Line Items sheet."""
    line_items: list[dict] = result.discovered.get("line_items") or []

    if not line_items:
        ws.cell(row=1, column=1, value="No line items")
        _autofit_columns(ws)
        return

    # Collect all column keys (union across all rows, preserving insertion order)
    all_keys: list[str] = []
    seen: set[str] = set()
    for item in line_items:
        for k in item.keys():
            if k not in seen:
                all_keys.append(k)
                seen.add(k)

    _write_header_row(ws, all_keys)
    ws.freeze_panes = "A2"

    for row_idx, item in enumerate(line_items, start=2):
        for col_idx, key in enumerate(all_keys, start=1):
            value = item.get(key)
            if _is_monetary_value(value) or _is_monetary_column(key):
                try:
                    numeric = _to_numeric(value) if value is not None else None
                    cell = ws.cell(row=row_idx, column=col_idx, value=numeric)
                    if numeric is not None:
                        cell.number_format = _MONEY_FMT
                except (TypeError, ValueError):
                    ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else None)
            else:
                ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else None)

    _autofit_columns(ws)


def _build_raw_text_sheet(ws: Any, result: ExtractionResult) -> None:
    """Populate the Raw Text sheet."""
    raw_text = result.discovered.get("raw_text", "") or ""
    ws.freeze_panes = ws["A2"]
    cell = ws.cell(row=1, column=1, value=raw_text)
    cell.alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 80
    ws.row_dimensions[1].height = max(15, min(400, raw_text.count("\n") * 15 + 15))


def _build_metadata_sheet(ws: Any, result: ExtractionResult, filename: str) -> None:
    """Populate the Metadata sheet."""
    _write_header_row(ws, ["Field", "Value"])
    ws.freeze_panes = "A2"

    issues = result.issues or []
    error_count = sum(1 for i in issues if i.severity == "error")

    meta_rows = [
        ("Filename", filename),
        ("Model", result.llm_model or ""),
        ("Extraction Timestamp", result.extraction_timestamp or ""),
        ("Requires Review", str(result.requires_review)),
        ("Issues (total)", str(len(issues))),
        ("Issues (errors)", str(error_count)),
    ]

    for row_idx, (field, value) in enumerate(meta_rows, start=2):
        ws.cell(row=row_idx, column=1, value=field)
        cell = ws.cell(row=row_idx, column=2, value=value)
        # Issues (errors) row: red font if any errors
        if field == "Issues (errors)" and error_count > 0:
            cell.font = _RED_FONT

    _autofit_columns(ws)


# ── Public API ─────────────────────────────────────────────────────────────────

def to_xlsx(
    result: ExtractionResult,
    filename: str,
    template_fields: list[dict] | None = None,
) -> bytes:
    """
    Export ExtractionResult to Excel bytes.

    4 sheets:
    1. Summary      — anchor fields (or template-filtered) as Field/Value rows
    2. Line Items   — discovered.line_items as a table (auto-detect columns)
    3. Raw Text     — single cell with discovered.raw_text
    4. Metadata     — extraction timestamp, model, review flag, issues count

    All sheets: freeze top row, auto-fit columns (min 10, max 50 chars).
    Summary + Metadata: bold blue (#2D5B8A) header row.
    Monetary values (Decimal or float): numeric with #,##0.00 format.
    """
    wb = openpyxl.Workbook()

    # openpyxl creates one default sheet — rename it to "Summary"
    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_lines = wb.create_sheet("Line Items")
    ws_raw = wb.create_sheet("Raw Text")
    ws_meta = wb.create_sheet("Metadata")

    _build_summary_sheet(ws_summary, result, template_fields)
    _build_line_items_sheet(ws_lines, result)
    _build_raw_text_sheet(ws_raw, result)
    _build_metadata_sheet(ws_meta, result, filename)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
