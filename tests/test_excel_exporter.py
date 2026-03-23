from __future__ import annotations

from decimal import Decimal
from io import BytesIO

import openpyxl
import pytest

from backend.schemas.extraction import AnchorFields, ExtractionIssue, ExtractionResult
from backend.services.excel_exporter import to_xlsx


def _make_result(
    line_items=None,
    issues=None,
    raw_text: str = "Sample raw text from OCR.",
) -> ExtractionResult:
    anchor = AnchorFields(
        issuer_name="INMOBILIARIA BARUAL S.L.",
        issuer_cif="B50042332",
        recipient_name="GESTINAD ARAGON S.L.",
        recipient_cif="B99334708",
        invoice_number="GAT143/26",
        issue_date="2026-03-01",
        base_imponible=Decimal("102.57"),
        iva_rate=Decimal("21"),
        iva_amount=Decimal("21.54"),
        irpf_rate=None,
        irpf_amount=None,
        total_amount=Decimal("124.11"),
        currency="EUR",
    )
    return ExtractionResult(
        anchor=anchor,
        discovered={
            "line_items": line_items
            if line_items is not None
            else [{"concept": "RENTA LEGAL", "total_amount": Decimal("102.57")}],
            "raw_text": raw_text,
        },
        issues=issues
        or [
            ExtractionIssue(
                field=None,
                message="IVA looks correct",
                severity="observation",
                source="llm",
            )
        ],
        requires_review=False,
        llm_model="qwen3.5:9b",
        extraction_timestamp="2026-03-01T10:00:00Z",
    )


# ── Test 1 ────────────────────────────────────────────────────────────────────
def test_to_xlsx_returns_bytes():
    result = _make_result()
    output = to_xlsx(result, "invoice.pdf")
    assert isinstance(output, bytes)
    assert len(output) > 0


# ── Test 2 ────────────────────────────────────────────────────────────────────
def test_to_xlsx_has_four_sheets():
    result = _make_result()
    output = to_xlsx(result, "invoice.pdf")
    wb = openpyxl.load_workbook(BytesIO(output))
    assert wb.sheetnames == ["Summary", "Line Items", "Raw Text", "Metadata"]


# ── Test 3 ────────────────────────────────────────────────────────────────────
def test_to_xlsx_summary_sheet_header():
    result = _make_result()
    output = to_xlsx(result, "invoice.pdf")
    wb = openpyxl.load_workbook(BytesIO(output))
    ws = wb["Summary"]
    row1 = [ws.cell(1, 1).value, ws.cell(1, 2).value]
    assert row1 == ["Field", "Value"]


# ── Test 4 ────────────────────────────────────────────────────────────────────
def test_to_xlsx_monetary_value_is_numeric():
    """Decimal anchor fields should land as numeric (float/int) cells, not strings."""
    result = _make_result()
    output = to_xlsx(result, "invoice.pdf")
    wb = openpyxl.load_workbook(BytesIO(output))
    ws = wb["Summary"]

    # Find the row whose Field cell is "base_imponible" and check Value type
    monetary_cell = None
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value == "base_imponible":
            monetary_cell = row[1]
            break

    assert monetary_cell is not None, "base_imponible row not found in Summary sheet"
    assert isinstance(
        monetary_cell.value, (int, float)
    ), f"Expected numeric, got {type(monetary_cell.value)}"


# ── Test 5 ────────────────────────────────────────────────────────────────────
def test_to_xlsx_with_template_filter():
    """With a template that includes only one field, Summary has exactly 2 rows (header + 1 data)."""
    result = _make_result()
    template_fields = [
        {
            "field_path": "anchor.invoice_number",
            "display_name": "Número de Factura",
            "include": True,
        },
        {
            "field_path": "anchor.issuer_name",
            "display_name": "Emisor",
            "include": False,  # excluded
        },
    ]
    output = to_xlsx(result, "invoice.pdf", template_fields=template_fields)
    wb = openpyxl.load_workbook(BytesIO(output))
    ws = wb["Summary"]
    # Count non-empty rows
    data_rows = [
        row for row in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in row)
    ]
    assert len(data_rows) == 1
    assert data_rows[0][0] == "Número de Factura"
    assert data_rows[0][1] == "GAT143/26"


# ── Test 6 ────────────────────────────────────────────────────────────────────
def test_to_xlsx_line_items_sheet():
    """Line Items sheet must have the right column headers when line_items are present."""
    line_items = [
        {"concept": "RENTA LEGAL", "quantity": 1, "unit_price": Decimal("102.57"), "total_amount": Decimal("102.57")},
        {"concept": "OTRO CONCEPTO", "quantity": 2, "unit_price": Decimal("50.00"), "total_amount": Decimal("100.00")},
    ]
    result = _make_result(line_items=line_items)
    output = to_xlsx(result, "invoice.pdf")
    wb = openpyxl.load_workbook(BytesIO(output))
    ws = wb["Line Items"]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    for key in ("concept", "quantity", "unit_price", "total_amount"):
        assert key in headers, f"Expected column '{key}' in Line Items headers: {headers}"
